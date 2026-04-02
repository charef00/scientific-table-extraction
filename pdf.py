import os
import cv2
import requests
import base64
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from pdf2image import convert_from_path
from ultralytics import YOLO



def pdf_to_tables_png(
    pdf_path,
    poppler_path=r"poppler-25.12.0\Library\bin",
    dpi=300,
    output_dir="tables",
    model_path="yolov8x-doclaynet.pt",
    conf=0.4,
    device="cpu"
):
    doi = os.path.splitext(os.path.basename(pdf_path))[0]
    os.makedirs(output_dir, exist_ok=True)

    # Convert PDF to images (PIL)
    pages = convert_from_path(
        pdf_path,
        dpi=dpi,
        poppler_path=poppler_path
    )

    if len(pages) <= 4:
        raise ValueError("PDF must contain more than 4 pages")

    # Remove first 2 and last 2 pages
    pages = pages[2:-2]

    # Load YOLO model once
    model = YOLO(model_path)

    table_count = 0

    for page_idx, page in enumerate(pages, start=1):

        # PIL → OpenCV
        page_cv = cv2.cvtColor(np.array(page), cv2.COLOR_RGB2BGR)

        # YOLO inference
        results = model(page_cv, conf=conf, device=device)
        r = results[0]

        for box, cls, score in zip(
            r.boxes.xyxy,
            r.boxes.cls,
            r.boxes.conf
        ):
            label = r.names[int(cls)]

            if label == "Table":
                x1, y1, x2, y2 = map(int, box)

                table_img = page_cv[y1:y2, x1:x2]

                table_count += 1
                table_name = f"{doi}_{table_count}.png"
                table_path = os.path.join(output_dir, table_name)

                cv2.imwrite(table_path, table_img)

    return doi, table_count

#create account in padlle /: is  free  and create api with token
def extract_table_html(file_path):
    api_url="api_url"
    token= "token"
    # Read file and encode to base64
    with open(file_path, "rb") as file:
        file_bytes = file.read()
        file_data = base64.b64encode(file_bytes).decode("ascii")

    headers = {
        "Authorization": f"token {token}",
        "Content-Type": "application/json"
    }

    payload = {
        "file": file_data,
        "fileType": 1,  # 0 = PDF, 1 = Image
        "useDocOrientationClassify": False,
        "useDocUnwarping": False,
        "useChartRecognition": False,
    }

    response = requests.post(api_url, json=payload, headers=headers)
    response.raise_for_status()

    result = response.json()["result"]
    table_html = result["layoutParsingResults"][0]["markdown"]["text"]

    return table_html

def html_table_to_excel(html_table: str, output_path: str = "output.xlsx"):
    
    soup = BeautifulSoup(html_table, "html.parser")
    table = soup.find("table")

    if table is None:
        raise ValueError("No <table> found in provided HTML")

    wb = Workbook()
    ws = wb.active

    occupied = set()
    row_idx = 1

    for row in table.find_all("tr"):
        col_idx = 1

        for cell in row.find_all(["td", "th"]):
            # Skip cells already covered by rowspan/colspan
            while (row_idx, col_idx) in occupied:
                col_idx += 1

            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            value = cell.get_text(strip=True)

            ws.cell(row=row_idx, column=col_idx, value=value)

            # Mark occupied cells
            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    occupied.add((r, c))

            # Merge cells if needed
            if rowspan > 1 or colspan > 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx + rowspan - 1,
                    end_column=col_idx + colspan - 1
                )

            col_idx += colspan

        row_idx += 1

    wb.save(output_path)

